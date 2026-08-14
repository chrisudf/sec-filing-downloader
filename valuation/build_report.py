# -*- coding: utf-8 -*-
"""通用估值报告生成器：engine.py 输出的 valuation JSON -> 六表 Excel 工作簿。

用法: python build_report.py VALUATION.json [OUT.xlsx]
默认输出到仓库 reports/{TICKER}_valuation_{date}.xlsx（reports/ 已在 .gitignore）。
工作表：摘要 / 情景假设 / DCF / SOTP / 历史数据 / 出处；黄色格为可调假设，改动后全表联动。
"""
import json
import os
import sys

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SP = os.path.dirname(os.path.abspath(__file__))
d = json.load(open(sys.argv[1], encoding="utf-8"))
T = d["ticker"]
OUT = sys.argv[2] if len(sys.argv) > 2 else os.path.join(
    os.path.dirname(SP), "reports", f"{T}_valuation_{d['date']}.xlsx")

BLUE = Font(name="Arial", color="0000FF", size=10)
BLACK = Font(name="Arial", color="000000", size=10)
GREEN = Font(name="Arial", color="008000", size=10)
BOLD = Font(name="Arial", bold=True, size=10)
TITLE = Font(name="Arial", bold=True, size=14)
H2 = Font(name="Arial", bold=True, size=11)
YELLOW = PatternFill("solid", fgColor="FFFF00")
HDRFILL = PatternFill("solid", fgColor="D9D9F3")
GREENFILL = PatternFill("solid", fgColor="E2EFDA")
REDFILL = PatternFill("solid", fgColor="FCE4E4")
THIN = Border(*[Side(style="thin", color="BFBFBF")] * 4)

FM_M = '$#,##0;($#,##0);"-"'
FM_PCT = '0.0%;(0.0%);"-"'
FM_X = '0.0"x"'
FM_PX = '$#,##0.0'
FM_EPS = '$0.00'

wb = openpyxl.Workbook()


def put(ws, cell, value, font=BLACK, fmt=None, fill=None, border=True, wrap=False):
    c = ws[cell]
    c.value = value
    c.font = font
    if fmt:
        c.number_format = fmt
    if fill:
        c.fill = fill
    if border:
        c.border = THIN
    if wrap:
        c.alignment = Alignment(wrap_text=True, vertical="top")
    return c


S = d["scenarios"]
R = d["rationale"]
FWD = d["meta"]["fwd_label"]
MODE = d.get("mode", "standard")

# ttm_revenue_override 生效时只有营收被滚动到含最新季度的窗口，同块的利润/现金流
# 仍是 XBRL 的上一窗口。不标窗口的话，"关键事实（勿改）"里会出现两类矛盾：
# 按行相除得到跨期利润率（SOFI 实测 646/4306=15.0%，分子缺最新季度利润）；
# 已滚动的调整后净利高于陈旧的报告净利（637 vs 577），而 adj_note 恰好说二者相等。
# 窗口日期必须取 data_latest（XBRL 数据期末），不能用 vintage.report_end（定期报告
# 期末）——TSM 这类 6-K 发行人二者分别是 2024-12-31 与 2026-06-30，标错等于把
# "严重滞后"写成"刚更新"。老 valuation.json 没这个键，缺了就不写日期。
_asof = d["meta"].get("data_latest")
STALE = (f"；仍为 XBRL 窗口（至 {_asof}），不含最新季度" if _asof
         else "；仍为 XBRL 窗口，不含最新季度") if d.get("rev0_note") else ""

if MODE == "financials":
    # ============================================================
    # 金融股工作簿：摘要 / 情景假设 / PTBV敏感性 / 历史数据 / 出处
    # 方法 = PE 法 + P/TBV 法（Gordon justified P/TBV 作参考列）
    # ============================================================
    FM_X2 = '0.00"x"'
    M = d["meta"]

    # ---------- 情景假设 ----------
    ws = wb.active
    ws.title = "情景假设"
    ws.column_dimensions["A"].width = 30
    for col in "BCD":
        ws.column_dimensions[col].width = 13
    ws.column_dimensions["E"].width = 66
    put(ws, "A1", f"{T} 情景假设与关键事实（金融股模式）", TITLE, border=False)
    put(ws, "A2", "黄色格 = 可调整的假设（改动后全表自动重算）；黑字 = 公式；出处见「出处」表",
        GREEN, border=False)
    for cell, v in [("A3", "参数"), ("B3", "悲观 Bear"), ("C3", "合理 Base"),
                    ("D3", "乐观 Bull"), ("E3", "依据（判断层，可人工复核）")]:
        put(ws, cell, v, BOLD, fill=HDRFILL)
    params = [
        (f"{FWD} 总净收入增速 (vs TTM)", "g", FM_PCT, R["g"]),
        (f"{FWD} 净利率（净利/总净收入）", "nm", FM_PCT, R["nm"]),
        (f"目标 PE (对 {FWD} EPS)", "pe", '0"x"', R["pe"]),
        ("目标 P/TBV", "ptbv", FM_X2, R["ptbv"]),
        ("股权资本成本 WACC", "wacc", FM_PCT, R["wacc"]),
        ("永续增长率", "tg", FM_PCT, "长期名义 GDP 附近；用于 justified P/TBV"),
    ]
    r = 4
    for label, key, fmt, note in params:   # 行号：g=4 nm=5 pe=6 ptbv=7 wacc=8 tg=9
        put(ws, f"A{r}", label, BOLD)
        for col, sc in zip("BCD", ("bear", "base", "bull")):
            put(ws, f"{col}{r}", S[sc]["assumptions"][key], BLUE, fmt=fmt, fill=YELLOW)
        put(ws, f"E{r}", note, GREEN)
        r += 1
    put(ws, "A11", "前瞻期(NTM)稀释股本 (M)", BOLD)
    put(ws, "B11", M["fwd_shares"], BLUE, fmt="#,##0", fill=YELLOW)
    put(ws, "E11", "三情景共用", GREEN)

    put(ws, "A13", "关键事实（SEC XBRL / 财报原文，勿改）", H2, border=False)
    facts_rows = [
        ("TTM 总净收入 ($M)", d["ttm"]["revenue"],
         d.get("rev0_note") or "含净利息收入的总净收入口径"),
        ("TTM 税前利润 ($M)", d["ttm"]["pretax_income"],
         "银行不申报营业利润，以税前利润代之" + STALE),
        ("TTM 报告净利 ($M)", d["ttm"]["net_income"], "含一次性项目的账面口径" + STALE),
        ("TTM 调整后净利 ($M)", d["adj_ni"], d["adj_note"]),
        ("有形账面价值 TBV ($M)", M["tbv"],
         f"股东权益−商誉−无形资产（XBRL {M['tbv_asof']} 时点）"),
        ("稀释股本 (M)", M["shares"], "最新期加权稀释股数"),
    ]
    r = 14   # 行号：rev=14 pretax=15 ni=16 adj=17 tbv=18 shares=19
    for label, v, note in facts_rows:
        put(ws, f"A{r}", label, BOLD)
        put(ws, f"B{r}", v, BLUE, fmt=FM_M if "股本" not in label else "#,##0")
        put(ws, f"E{r}", note, GREEN, wrap=True)
        r += 1
    put(ws, "A20", "TTM 调整后 EPS", BOLD)
    put(ws, "B20", "=B17/B19", BLACK, fmt=FM_EPS)
    put(ws, "A21", "每股 TBV", BOLD)
    put(ws, "B21", "=B18/B19", BLACK, fmt=FM_PX)
    put(ws, "A22", "TTM ROTE（调整后）", BOLD)
    put(ws, "B22", "=B17/B18", BLACK, fmt=FM_PCT)

    put(ws, "A24", "情景计算（公式）", H2, border=False)
    calc_rows = [
        (f"{FWD} 总净收入 ($M)", "=$B$14*(1+{c}4)", FM_M),
        (f"{FWD} 净利 ($M)", "={c}25*{c}5", FM_M),
        (f"{FWD} EPS", "={c}26/$B$11", FM_EPS),
        ("PE 法目标价", "={c}27*{c}6", FM_PX),
        (f"{FWD} ROTE", "={c}26/$B$18", FM_PCT),
        ("justified P/TBV = (ROTE−g)/(WACC−g)", "=({c}29-{c}9)/({c}8-{c}9)", FM_X2),
        ("P/TBV 法目标价", "=$B$21*{c}7", FM_PX),
        ("justified 参考价", "=$B$21*{c}30", FM_PX),
    ]
    for j, (label, f, fmt) in enumerate(calc_rows):   # 行号 25-32
        rr = 25 + j
        put(ws, f"A{rr}", label, BOLD)
        for col in "BCD":
            put(ws, f"{col}{rr}", f.format(c=col), BLACK, fmt=fmt)

    # ---------- PTBV 敏感性 ----------
    ws = wb.create_sheet("PTBV敏感性")
    ws.column_dimensions["A"].width = 16
    for col in "BCDEF":
        ws.column_dimensions[col].width = 12
    put(ws, "A1", f"{T} justified P/TBV 参考价 = f(WACC, 永续g)（合理情景 ROTE）", TITLE, border=False)
    put(ws, "A2", "注：由 Python 估值引擎离线计算（每股 TBV × (ROTE−g)/(WACC−g)）；"
                  "改假设后以「情景假设」公式区为准", GREEN, border=False)
    sens = d["sensitivity"]
    # JSON 键是字符串，负 g（低永续增长情景）下字典序 ≠ 数值序，必须按 float 排
    gs = sorted({g for row in sens.values() for g in row}, key=float)
    put(ws, "A4", "WACC \\ 永续g", BOLD, fill=HDRFILL)
    for j, g in enumerate(gs):
        put(ws, f"{'BCDEF'[j]}4", float(g), BOLD, fmt=FM_PCT, fill=HDRFILL)
    for i, (w, row) in enumerate(sorted(sens.items(), key=lambda kv: float(kv[0]))):
        put(ws, f"A{5+i}", float(w), BOLD, fmt=FM_PCT, fill=HDRFILL)
        for j, g in enumerate(gs):
            put(ws, f"{'BCDEF'[j]}{5+i}", row[g], BLUE, fmt=FM_PX)

    # ---------- 历史数据 ----------
    ws = wb.create_sheet("历史数据")
    ws.column_dimensions["A"].width = 22
    for i in range(2, 11):
        ws.column_dimensions[get_column_letter(i)].width = 12
    put(ws, "A1", f"{T} 历史财务数据（SEC XBRL companyfacts）", TITLE, border=False)
    ha = d["history"]["annual"]
    n = len(ha["labels"])
    put(ws, "A3", "财年(期末) ($M)", BOLD, fill=HDRFILL)
    for j, lab in enumerate(ha["labels"]):
        put(ws, f"{get_column_letter(2+j)}3", lab, BOLD, fill=HDRFILL)
    for r_i, (label, key, fmt) in enumerate(
            [("总净收入", "rev", FM_M), ("税前利润", "op", FM_M),
             ("净利润", "ni", FM_M), ("稀释 EPS", "eps", FM_EPS)], start=4):
        put(ws, f"A{r_i}", label, BOLD)
        for j in range(n):
            v = ha[key][j]
            put(ws, f"{get_column_letter(2+j)}{r_i}", v if v is not None else "-", BLUE, fmt=fmt)
    put(ws, "A8", "净利率", BOLD)
    put(ws, "A9", "总净收入 YoY", BOLD)
    for j in range(n):
        col = get_column_letter(2 + j)
        put(ws, f"{col}8", f"={col}6/{col}4", BLACK, fmt=FM_PCT)
        if j > 0:
            put(ws, f"{col}9", f"={col}4/{get_column_letter(1+j)}4-1", BLACK, fmt=FM_PCT)
    hq = d["history"]["quarterly"]
    nq = len(hq["labels"])
    put(ws, "A12", "近八个季度 ($M)", BOLD, fill=HDRFILL)
    for j, lab in enumerate(hq["labels"]):
        put(ws, f"{get_column_letter(2+j)}12", lab, BOLD, fill=HDRFILL)
    for r_i, (label, key) in enumerate(
            [("总净收入", "rev"), ("税前利润", "op"), ("净利润", "ni")], start=13):
        put(ws, f"A{r_i}", label, BOLD)
        for j in range(nq):
            put(ws, f"{get_column_letter(2+j)}{r_i}", hq[key][j], BLUE, fmt=FM_M)
    put(ws, "A16", "净利率", BOLD)
    for j in range(nq):
        col = get_column_letter(2 + j)
        put(ws, f"{col}16", f"={col}15/{col}13", BLACK, fmt=FM_PCT)
    put(ws, "A18", "注：Q4 为年度减前三季推算；净利率波动大的季度多为拨备/一次性项目，见「摘要」注记。",
        GREEN, border=False)

    # ---------- 摘要 ----------
    ws = wb.create_sheet("摘要", 0)
    ws.column_dimensions["A"].width = 27
    for col in "BCDEF":
        ws.column_dimensions[col].width = 14
    ws.column_dimensions["H"].width = 72
    put(ws, "A1", f"{d['name']} / {T} 估值分析（金融股模式）", TITLE, border=False)
    put(ws, "A2", f"sec-filing-downloader + SEC XBRL · {d['date']} · 分析工具输出，不构成投资建议",
        GREEN, border=False)
    caliber = ("方法：PE 法 + P/TBV 法（金融股不适用 FCFF DCF）"
               "；口径注意：金融股模式沿用 v1 情景语义（情景盈利×情景倍数，无 v2 一致性联动）")
    if M.get("adr_multiple", 1.0) != 1.0:
        caliber += f"；价格为 ADR（1 ADR = {M['adr_multiple']:g} 普通股，股本已折算）"
    if M.get("currency", "USD") != "USD":
        caliber += f"；申报货币 {M['currency']}，已按现汇折算美元"
    put(ws, "A3", caliber, GREEN, border=False)

    put(ws, "A4", "现价（黄色可改，全表联动）", BOLD)
    put(ws, "B4", M["price"], BLUE, fmt=FM_PX, fill=YELLOW)
    put(ws, "A5", "市值 ($M)", BOLD)
    put(ws, "B5", "=B4*'情景假设'!$B$19", BLACK, fmt=FM_M)
    put(ws, "A6", "每股 TBV", BOLD)
    put(ws, "B6", "='情景假设'!$B$21", GREEN, fmt=FM_PX)
    put(ws, "A7", "TTM ROTE（调整后）", BOLD)
    put(ws, "B7", "='情景假设'!$B$22", GREEN, fmt=FM_PCT)

    _vin = M.get("vintage") or {}
    if _vin:
        put(ws, "D4", "数据时效（价值投资第一检查项）", BOLD, fill=HDRFILL)
        put(ws, "E4", "", BOLD, fill=HDRFILL)
        put(ws, "D5", "最新报告期末", BOLD)
        put(ws, "E5", _vin["report_end"], BLACK)
        put(ws, "D6", "提交于", BOLD)
        put(ws, "E6", _vin["filed"], BLACK)
        put(ws, "D7", "距报告期(天)", BOLD)
        _c = put(ws, "E7", _vin["age_days"], BLACK, fmt="#,##0")
        if _vin["age_days"] > 120:
            _c.font = Font(name="Arial", color="CC0000", bold=True)
        if _vin.get("pending_8k_announced"):
            _c = put(ws, "D8", f"⚠ 业绩已公布({_vin['pending_8k_announced']} 8-K)", BOLD)
            _c.font = Font(name="Arial", color="CC0000", bold=True)
            _c = put(ws, "E8", "10-Q 未提交:以上定量指标仍基于上一报告期", BOLD)
            _c.font = Font(name="Arial", color="CC0000", bold=True)

    put(ws, "A9", "关键指标", H2, border=False)
    for cell, v in [("A10", "指标"), ("B10", "数值"), ("H10", "说明")]:
        put(ws, cell, v, BOLD, fill=HDRFILL)
    metrics = [
        ("TTM 调整后 EPS", "='情景假设'!$B$20", FM_EPS, d["adj_note"]),
        ("调整后 Trailing PE", "=B4/B11", FM_X, "现价 / TTM 调整后 EPS"),
        (f"{FWD} EPS（合理）", "='情景假设'!$C$27", FM_EPS, "合理情景假设下的前瞻期(NTM) EPS"),
        ("Forward PE（合理）", "=B4/B13", FM_X, ""),
        ("当前 P/TBV", "=B4/'情景假设'!$B$21", FM_X2, "现价 / 每股有形账面价值"),
        ("justified P/TBV（合理）", "='情景假设'!$C$30", FM_X2,
         "(前瞻ROTE−永续g)/(WACC−永续g)，Gordon 公式交叉参考"),
    ]
    r = 11
    for label, v, fmt, note in metrics:
        put(ws, f"A{r}", label, BOLD)
        put(ws, f"B{r}", v, BLACK, fmt=fmt)
        put(ws, f"H{r}", note, GREEN, wrap=True)
        r += 1

    put(ws, "A19", "投资情景（两种方法 + 综合）", H2, border=False)
    for j, h in enumerate(["投资情景", "PE 法目标价", "P/TBV 法目标价",
                           "综合目标价", "距现价", "一年后目标价"]):
        put(ws, f"{get_column_letter(1+j)}20", h, BOLD, fill=HDRFILL)
    scen_rows = [("🚀 乐观 (Bull)", "D"), ("⚖️ 合理 (Base)", "C"), ("🛡️ 悲观 (Bear)", "B")]
    r = 21   # 行号：bull=21 base=22 bear=23
    for label, ac in scen_rows:
        put(ws, f"A{r}", label, BOLD)
        put(ws, f"B{r}", f"='情景假设'!${ac}$28", GREEN, fmt=FM_PX)
        put(ws, f"C{r}", f"='情景假设'!${ac}$31", GREEN, fmt=FM_PX)
        c = put(ws, f"D{r}", f"=AVERAGE(B{r}:C{r})", BLACK, fmt=FM_PX)
        c.fill = GREENFILL if r != 23 else REDFILL
        c.font = Font(name="Arial", bold=True, size=10)
        put(ws, f"E{r}", f"=D{r}/$B$4-1", BLACK, fmt=FM_PCT)
        put(ws, f"F{r}", f"=D{r}*(1+'情景假设'!${ac}$8)", BLACK, fmt=FM_PX)
        r += 1
    from openpyxl.formatting.rule import CellIsRule  # noqa: E402
    ws.conditional_formatting.add(
        "E21:E23", CellIsRule(operator="lessThan", formula=["0"],
                              font=Font(name="Arial", color="CC0000", bold=True)))
    ws.conditional_formatting.add(
        "E21:E23", CellIsRule(operator="greaterThan", formula=["0"],
                              font=Font(name="Arial", color="008000", bold=True)))
    put(ws, "A24", "注：一年后目标价 = 综合目标价 × (1+该情景 WACC)；justified P/TBV 为交叉参考不入综合",
        GREEN, border=False)
    _spread = " / ".join(f"{sc} {d['scenarios'][sc].get('method_spread') or '—'}x"
                         for sc in ("bear", "base", "bull"))
    put(ws, "A25", f"方法离散度（PE法 vs P/TBV法 max/min）：{_spread}——离散大时两法分歧大，"
                   "综合可信度降低", GREEN, border=False)
    put(ws, "A26", "判断层注记（均有财报原文出处）：", H2, border=False)
    for j, note in enumerate(d["notes"]):
        put(ws, f"A{27+j}", note, GREEN, border=False)

    # ---------- 出处 ----------
    ws = wb.create_sheet("出处")
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 105
    put(ws, "A1", "数据出处与调整口径", TITLE, border=False)
    sources = [
        ("下载器 manifest", d["manifest"].strip()),
        ("XBRL companyfacts", f"https://data.sec.gov/api/xbrl/companyfacts/CIK*.json（{T}）"
                              "—— 总净收入/税前利润/净利/EPS/股东权益/商誉/无形资产/股本"),
        ("现价", f"yfinance {T} fast_info，{d['date']}：${M['price']}，市值 ${M['mcap']:,}M"),
        ("TTM 口径", "损益类=最近四个离散季度加总（Q4=年度-前三季）；无季度 XBRL 时退回最新财年"),
        ("调整后净利", d["adj_note"]),
        ("TBV 口径", f"股东权益−商誉−无形资产，XBRL {M['tbv_asof']} 时点"),
        ("估值方法", "PE 法 + P/TBV 法（银行/券商 CFO 含贷款投放，FCFF DCF 不适用）；"
                    "justified P/TBV=(ROTE−g)/(WACC−g) 仅作交叉参考"),
        ("免责声明", "本报告为 sec-filing-downloader 项目的数据分析输出，"
                    "所有估值结果由假设驱动，不构成任何投资建议。"),
    ]
    r = 3
    for label, note in sources:
        put(ws, f"A{r}", label, BOLD)
        put(ws, f"B{r}", note, BLACK, wrap=True)
        r += 1

    os.makedirs(os.path.dirname(OUT) or ".", exist_ok=True)
    wb.save(OUT)
    print("saved:", OUT)
    sys.exit(0)

# ============================================================ 情景假设
ws = wb.active
ws.title = "情景假设"
ws.column_dimensions["A"].width = 30
for col in "BCD":
    ws.column_dimensions[col].width = 13
ws.column_dimensions["E"].width = 66

put(ws, "A1", f"{T} 情景假设与关键事实", TITLE, border=False)
put(ws, "A2", "黄色格 = 可调整的假设（改动后全表自动重算）；黑字 = 公式；出处见「出处」表", GREEN, border=False)
for cell, v in [("A3", "参数"), ("B3", "悲观 Bear"), ("C3", "合理 Base"), ("D3", "乐观 Bull"), ("E3", "依据（判断层，可人工复核）")]:
    put(ws, cell, v, BOLD, fill=HDRFILL)

params = [
    (f"{FWD} 营收增速 (vs TTM)", "g", FM_PCT, R["g"]),
    (f"{FWD} 营业利润率", "opm", FM_PCT, R["opm"]),
    ("正常化税率", "tax", FM_PCT, d["adj_note"][:40] + "…"),
    (f"目标 PE (对 {FWD} EPS)", "pe", '0"x"', R["pe"]),
    ("主分部 EV/EBIT 倍数", "m1", '0"x"', R["m1"]),
    ("次分部 EV/EBIT 倍数", "m2", '0"x"', R["rl"]),
    ("WACC", "wacc", FM_PCT, R["wacc"]),
    ("永续增长率", "tg", FM_PCT, "长期名义 GDP 附近"),
    ("DCF 起始增速 (第1年)", "g0", FM_PCT, "十年线性衰减起点，FCF利润率路径见 DCF 表"),
    ("DCF 第10年增速", "gN", FM_PCT, "线性衰减终点"),
]
r = 4
for label, key, fmt, note in params:
    put(ws, f"A{r}", label, BOLD)
    for col, sc in zip("BCD", ("bear", "base", "bull")):
        put(ws, f"{col}{r}", S[sc]["assumptions"][key], BLUE, fmt=fmt, fill=YELLOW)
    put(ws, f"E{r}", note, GREEN)
    r += 1
put(ws, "A14", "前瞻期(NTM)稀释股本 (M)", BOLD)
put(ws, "B14", d["meta"]["fwd_shares"], BLUE, fmt="#,##0", fill=YELLOW)
put(ws, "E14", "回购小幅缩减；三情景共用", GREEN)
put(ws, "A15", "年化其他收益 ($M)", BOLD)
put(ws, "B15", d["other_income"], BLUE, fmt=FM_M, fill=YELLOW)
put(ws, "E15", "净利息及其他（正常化）；三情景共用", GREEN)
put(ws, "A16", "主分部利润占比", BOLD)
put(ws, "B16", d["meta"]["seg1_share"], BLUE, fmt=FM_PCT, fill=YELLOW)
put(ws, "E16", f"主分部={d['meta']['seg1']}；次分部={d['meta']['seg2']}", GREEN)

put(ws, "A18", "关键事实（SEC XBRL / 财报原文，勿改）", H2, border=False)
facts_rows = [
    ("TTM 营收 ($M)", d["ttm"]["revenue"],
     d.get("rev0_note") or "XBRL 最近四个离散季度加总"),
    # override 生效时上一行的说明变成了营收的原文出处，"同上"会顺带宣称营业利润
    # 也来自滚动后的新闻稿——那时必须写回 XBRL 口径并标窗口
    ("TTM 营业利润 ($M)", d["ttm"]["op_income"],
     ("XBRL 最近四个离散季度加总" + STALE) if STALE else "同上"),
    ("TTM 报告净利 ($M)", d["ttm"]["net_income"], "含一次性项目的账面口径" + STALE),
    ("TTM 调整后净利 ($M)", d["adj_ni"], d["adj_note"]),
    ("TTM 经营现金流 ($M)", d["ttm"]["cfo"],
     "年度 + YTD 差额口径（10-Q现金流表为累计值）" + STALE),
    ("TTM 资本开支 ($M)", d["ttm"]["capex"], "同上"),
    (None, None, "CFO - Capex"),
    ("净现金 ($M)", d["meta"]["net_cash"], d["net_cash_note"]),
    ("稀释股本 (M)", d["meta"]["shares"], "最新季度加权稀释股数"),
]
r = 19
for label, v, note in facts_rows:
    if label is None:
        put(ws, f"A{r}", "TTM 自由现金流 ($M)", BOLD)
        put(ws, f"B{r}", "=B23-B24", BLACK, fmt=FM_M)
    else:
        put(ws, f"A{r}", label, BOLD)
        put(ws, f"B{r}", v, BLUE, fmt=FM_M if "股本" not in label else "#,##0")
    put(ws, f"E{r}", note, GREEN, wrap=True)
    r += 1
put(ws, "A28", "TTM 调整后 EPS", BOLD)
put(ws, "B28", "=B22/B27", BLACK, fmt=FM_EPS)
put(ws, "E28", "调整后净利 / 稀释股本", GREEN)

put(ws, "A30", "情景计算（公式）", H2, border=False)
calc_rows = [
    (f"{FWD} 营收 ($M)", "=$B$19*(1+{c}4)", FM_M),
    (f"{FWD} 营业利润 ($M)", "={c}31*{c}5", FM_M),
    (f"{FWD} 净利 ($M)", "=({c}32+$B$15)*(1-{c}6)", FM_M),
    (f"{FWD} EPS", "={c}33/$B$14", FM_EPS),
    ("PE 法目标价", "={c}34*{c}7", FM_PX),
]
for j, (label, f, fmt) in enumerate(calc_rows):
    rr = 31 + j
    put(ws, f"A{rr}", label, BOLD)
    for col in "BCD":
        put(ws, f"{col}{rr}", f.format(c=col), BLACK, fmt=fmt)

# ============================================================ DCF
ws = wb.create_sheet("DCF")
ws.column_dimensions["A"].width = 24
for col in "BCDEFGHIJK":
    ws.column_dimensions[col].width = 11
put(ws, "A1", f"{T} 十年两段式 FCFF DCF（基期=TTM营收；FCF利润率为可调输入）", TITLE, border=False)

COLS = "BCDEFGHIJK"
SCEN = [("悲观 Bear", "B", "bear", 3), ("合理 Base", "C", "base", 19), ("乐观 Bull", "D", "bull", 35)]
Y0 = 2027  # 第1年标签（NVDA FY28≈CY27 之类的差异在标签行说明，用 t+1..t+10 更通用）
for name, acol, key, r0 in SCEN:
    put(ws, f"A{r0}", f"—— {name} ——", H2, fill=HDRFILL)
    labels = ["年份 t+" + str(i + 1) for i in range(10)]
    put(ws, f"A{r0+1}", "预测期", BOLD)
    put(ws, f"A{r0+2}", "营收增速", BOLD)
    put(ws, f"A{r0+3}", "营收 ($M)", BOLD)
    put(ws, f"A{r0+4}", "FCF 利润率（可调）", BOLD)
    put(ws, f"A{r0+5}", "FCF ($M)", BOLD)
    put(ws, f"A{r0+6}", "折现值 ($M)", BOLD)
    for i, col in enumerate(COLS):
        put(ws, f"{col}{r0+1}", labels[i], BOLD, fill=HDRFILL)
        put(ws, f"{col}{r0+2}",
            f"='情景假设'!{acol}$12+('情景假设'!{acol}$13-'情景假设'!{acol}$12)*{i}/9",
            BLACK, fmt=FM_PCT)
        if i == 0:
            put(ws, f"{col}{r0+3}", f"='情景假设'!$B$19*(1+{col}{r0+2})", BLACK, fmt=FM_M)
        else:
            put(ws, f"{col}{r0+3}", f"={COLS[i-1]}{r0+3}*(1+{col}{r0+2})", BLACK, fmt=FM_M)
        put(ws, f"{col}{r0+4}", S[key]["assumptions"]["margins"][i], BLUE, fmt=FM_PCT, fill=YELLOW)
        put(ws, f"{col}{r0+5}", f"={col}{r0+3}*{col}{r0+4}", BLACK, fmt=FM_M)
        put(ws, f"{col}{r0+6}", f"={col}{r0+5}/POWER(1+'情景假设'!{acol}$10,{i+1})", BLACK, fmt=FM_M)
    tail = [
        ("显性期 PV 合计 ($M)", f"=SUM(B{r0+6}:K{r0+6})", FM_M),
        ("终值 ($M)", f"=K{r0+5}*(1+'情景假设'!{acol}11)/('情景假设'!{acol}10-'情景假设'!{acol}11)", FM_M),
        ("终值折现 ($M)", f"=B{r0+8}/POWER(1+'情景假设'!{acol}10,10)", FM_M),
        ("企业价值 EV ($M)", f"=B{r0+7}+B{r0+9}", FM_M),
        ("加：净现金 ($M)", "='情景假设'!$B$26", FM_M),
        ("股权价值 ($M)", f"=B{r0+10}+B{r0+11}", FM_M),
        ("每股价值", f"=B{r0+12}/'情景假设'!$B$27", FM_PX),
    ]
    for j, (label, formula, fmt) in enumerate(tail):
        put(ws, f"A{r0+7+j}", label, BOLD)
        c = put(ws, f"B{r0+7+j}", formula, BLACK, fmt=fmt)
        if label == "每股价值":
            c.fill = GREENFILL
            c.font = Font(name="Arial", bold=True, size=10)

put(ws, "A52", "敏感性分析（合理情景）：每股价值 = f(WACC, 永续增长率)", H2, border=False)
put(ws, "A53", "注：由 Python 估值引擎离线计算（同一公式）；改假设后以上方公式区为准", GREEN, border=False)
sens = d["sensitivity"]
# 同金融股分支：JSON 字符串键按 float 排，防负 g 时字典序错位
gs = sorted({g for row in sens.values() for g in row}, key=float)
put(ws, "A54", "WACC \\ 永续g", BOLD, fill=HDRFILL)
for j, g in enumerate(gs):
    put(ws, f"{'BCDEF'[j]}54", float(g), BOLD, fmt=FM_PCT, fill=HDRFILL)
for i, (w, row) in enumerate(sorted(sens.items(), key=lambda kv: float(kv[0]))):
    put(ws, f"A{55+i}", float(w), BOLD, fmt=FM_PCT, fill=HDRFILL)
    for j, g in enumerate(gs):
        put(ws, f"{'BCDEF'[j]}{55+i}", row[g], BLUE, fmt=FM_PX)

# ============================================================ SOTP
ws = wb.create_sheet("SOTP")
ws.column_dimensions["A"].width = 30
ws.column_dimensions["B"].width = 46
for col in "CDE":
    ws.column_dimensions[col].width = 14
_sotp_title = f"{T} 分部加总 / 倍数法（基于 {FWD} 营业利润）"
if not d["meta"].get("sotp_in_blend", True):
    _sotp_title += " ——参考项，不入综合（主分部利润占比>=85%，与 PE 法为同一笔盈利）"
put(ws, "A1", _sotp_title, TITLE, border=False)
for cell, v in [("A3", "项目"), ("B3", "说明"), ("C3", "悲观"), ("D3", "合理"), ("E3", "乐观")]:
    put(ws, cell, v, BOLD, fill=HDRFILL)

AC = {"C": "B", "D": "C", "E": "D"}
rows = [
    (f"主分部营业利润 ($M)", d["meta"]["seg1"], "='情景假设'!{a}32*'情景假设'!$B$16", FM_M),
    ("主分部 EV/EBIT 倍数", "引用「情景假设」第8行", "='情景假设'!{a}8", '0"x"'),
    ("主分部估值 ($M)", "", "={c}4*{c}5", FM_M),
    ("次分部营业利润 ($M)", d["meta"]["seg2"], "='情景假设'!{a}32*(1-'情景假设'!$B$16)", FM_M),
    ("次分部倍数", "引用「情景假设」第9行", "='情景假设'!{a}9", '0"x"'),
    ("次分部估值 ($M)", "", "={c}7*{c}8", FM_M),
    ("净现金 ($M)", d["net_cash_note"][:44] + "…", "='情景假设'!$B$26", FM_M),
    ("总权益价值 ($M)", "", "={c}6+{c}9+{c}10", FM_M),
    ("每股公允价值", "", "={c}11/'情景假设'!$B$27", FM_PX),
    ("现价", "可在「摘要」B4 修改", "='摘要'!$B$4", FM_PX),
    ("潜在收益", "", "={c}12/{c}13-1", FM_PCT),
]
for j, (label, desc, f, fmt) in enumerate(rows):
    rr = 4 + j
    put(ws, f"A{rr}", label, BOLD)
    put(ws, f"B{rr}", desc, GREEN)
    for col in "CDE":
        c = put(ws, f"{col}{rr}", f.format(a=AC[col], c=col), BLACK, fmt=fmt)
        if label == "每股公允价值":
            c.fill = GREENFILL
            c.font = Font(name="Arial", bold=True, size=10)

# ============================================================ 历史数据
ws = wb.create_sheet("历史数据")
ws.column_dimensions["A"].width = 22
for i in range(2, 11):
    ws.column_dimensions[get_column_letter(i)].width = 12
put(ws, "A1", f"{T} 历史财务数据（SEC XBRL companyfacts）", TITLE, border=False)

ha = d["history"]["annual"]
n = len(ha["labels"])
put(ws, "A3", "财年(期末) ($M)", BOLD, fill=HDRFILL)
for j, lab in enumerate(ha["labels"]):
    put(ws, f"{get_column_letter(2+j)}3", lab, BOLD, fill=HDRFILL)
for r_i, (label, key, fmt) in enumerate([("营收", "rev", FM_M), ("营业利润", "op", FM_M),
                                         ("净利润", "ni", FM_M), ("稀释 EPS", "eps", FM_EPS),
                                         ("经营现金流", "cfo", FM_M), ("资本开支", "capex", FM_M)], start=4):
    put(ws, f"A{r_i}", label, BOLD)
    for j in range(n):
        v = ha[key][j]
        put(ws, f"{get_column_letter(2+j)}{r_i}", v if v is not None else "-", BLUE, fmt=fmt)
put(ws, "A10", "自由现金流", BOLD)
put(ws, "A11", "营收 YoY", BOLD)
put(ws, "A12", "营业利润率", BOLD)
for j in range(n):
    col = get_column_letter(2 + j)
    put(ws, f"{col}10", f"={col}8-{col}9", BLACK, fmt=FM_M)
    if j > 0:
        put(ws, f"{col}11", f"={col}4/{get_column_letter(1+j)}4-1", BLACK, fmt=FM_PCT)
    put(ws, f"{col}12", f"={col}5/{col}4", BLACK, fmt=FM_PCT)

hq = d["history"]["quarterly"]
nq = len(hq["labels"])
put(ws, "A15", "近八个季度 ($M)", BOLD, fill=HDRFILL)
for j, lab in enumerate(hq["labels"]):
    put(ws, f"{get_column_letter(2+j)}15", lab, BOLD, fill=HDRFILL)
for r_i, (label, key) in enumerate([("营收", "rev"), ("营业利润", "op"), ("净利润", "ni")], start=16):
    put(ws, f"A{r_i}", label, BOLD)
    for j in range(nq):
        put(ws, f"{get_column_letter(2+j)}{r_i}", hq[key][j], BLUE, fmt=FM_M)
put(ws, "A19", "营业利润率", BOLD)
for j in range(nq):
    col = get_column_letter(2 + j)
    put(ws, f"{col}19", f"={col}17/{col}16", BLACK, fmt=FM_PCT)
put(ws, "A21", "注：期末日为各自财年/财季末；Q4 为年度减前三季推算。异常项见「摘要」注记。", GREEN, border=False)

# ============================================================ 摘要
ws = wb.create_sheet("摘要", 0)
ws.column_dimensions["A"].width = 27
for col in "BCDEFG":
    ws.column_dimensions[col].width = 14
ws.column_dimensions["H"].width = 72
put(ws, "A1", f"{d['name']} / {T} 估值分析", TITLE, border=False)
put(ws, "A2", f"sec-filing-downloader + SEC XBRL · {d['date']} · 分析工具输出，不构成投资建议", GREEN, border=False)
_caliber = "目标价口径：综合目标价 = 当前公允价值（锚定前瞻期 NTM 盈利与 DCF 现值，窗口见情景假设表）"
if d["meta"].get("adr_multiple", 1.0) != 1.0:
    _caliber += f"；价格为 ADR（1 ADR = {d['meta']['adr_multiple']:g} 普通股，股本已折算）"
if d["meta"].get("currency", "USD") != "USD":
    _caliber += f"；申报货币 {d['meta']['currency']}，已按现汇折算美元"
put(ws, "A3", _caliber, GREEN, border=False)

put(ws, "A4", "现价（黄色可改，全表联动）", BOLD)
put(ws, "B4", d["meta"]["price"], BLUE, fmt=FM_PX, fill=YELLOW)
put(ws, "A5", "市值 ($M)", BOLD)
put(ws, "B5", "=B4*'情景假设'!$B$27", BLACK, fmt=FM_M)
put(ws, "A6", "净现金 ($M)", BOLD)
put(ws, "B6", "='情景假设'!$B$26", GREEN, fmt=FM_M)
put(ws, "A7", "稀释股本 (M)", BOLD)
put(ws, "B7", "='情景假设'!$B$27", GREEN, fmt="#,##0")

_vin = d["meta"].get("vintage") or {}
if _vin:
    put(ws, "D4", "数据时效（价值投资第一检查项）", BOLD, fill=HDRFILL)
    put(ws, "E4", "", BOLD, fill=HDRFILL)
    put(ws, "D5", "最新报告期末", BOLD)
    put(ws, "E5", _vin["report_end"], BLACK)
    put(ws, "D6", "提交于", BOLD)
    put(ws, "E6", _vin["filed"], BLACK)
    put(ws, "D7", "距报告期(天)", BOLD)
    _c = put(ws, "E7", _vin["age_days"], BLACK, fmt="#,##0")
    if _vin["age_days"] > 120:
        _c.font = Font(name="Arial", color="CC0000", bold=True)
    if _vin.get("pending_8k_announced"):
        _c = put(ws, "D8", f"⚠ 业绩已公布({_vin['pending_8k_announced']} 8-K)", BOLD)
        _c.font = Font(name="Arial", color="CC0000", bold=True)
        _c = put(ws, "E8", "10-Q 未提交:以上定量指标仍基于上一报告期", BOLD)
        _c.font = Font(name="Arial", color="CC0000", bold=True)

put(ws, "A9", "关键指标", H2, border=False)
for cell, v in [("A10", "指标"), ("B10", "数值"), ("H10", "说明")]:
    put(ws, cell, v, BOLD, fill=HDRFILL)
metrics = [
    ("TTM 调整后 EPS", "='情景假设'!$B$28", FM_EPS, d["adj_note"]),
    ("调整后 Trailing PE", "=B4/B11", FM_X, "现价 / TTM 调整后 EPS"),
    (f"{FWD} EPS（合理）", "='情景假设'!$C$34", FM_EPS, "合理情景假设下的前瞻期(NTM) EPS"),
    ("Forward PE（合理）", "=B4/B13", FM_X, ""),
    ("PEG（合理）", "=B14/('情景假设'!$C$4*100)", "0.00", "Forward PE / 营收增速"),
    ("EV/EBIT (TTM)", "=(B5-B6)/'情景假设'!$B$20", FM_X, ""),
    ("FCF 收益率 (TTM)", "='情景假设'!$B$25/B5", FM_PCT, ""),
    ("反向 DCF 隐含增速（base 口径条件）", d["reverse_dcf"], FM_PCT,
     "现价隐含的 DCF 起始增速（10年线性衰减）——条件于 base 的利润率路径与 WACC，"
     "base 假设错则此值同错；Python 引擎计算"),
]
r = 11
for label, v, fmt, note in metrics:
    put(ws, f"A{r}", label, BOLD)
    put(ws, f"B{r}", v, BLUE if isinstance(v, float) else BLACK, fmt=fmt)
    put(ws, f"H{r}", note, GREEN, wrap=True)
    r += 1

# v2：主分部利润占比 >= 0.85 时 SOTP 只作参考（与 PE 法同一笔盈利），综合 = PE/DCF 均值
_sotp_in = d["meta"].get("sotp_in_blend", True)
put(ws, "A20", "投资情景（三种方法 + 综合）" if _sotp_in
    else "投资情景（PE/DCF 综合；SOTP 为参考）", H2, border=False)
for j, h in enumerate(["投资情景", "PE 法目标价", "DCF 每股价值",
                       "SOTP 每股价值" if _sotp_in else "SOTP（参考，不入综合）",
                       "综合目标价", "距现价", "一年后目标价"]):
    put(ws, f"{get_column_letter(1+j)}21", h, BOLD, fill=HDRFILL)
scen_rows = [
    ("🚀 乐观 (Bull)", "D", "=DCF!$B$48", "=SOTP!$E$12"),
    ("⚖️ 合理 (Base)", "C", "=DCF!$B$32", "=SOTP!$D$12"),
    ("🛡️ 悲观 (Bear)", "B", "=DCF!$B$16", "=SOTP!$C$12"),
]
_blend_rng = "B{r}:D{r}" if _sotp_in else "B{r}:C{r}"
r = 22
for label, ac, dcf_f, sotp_f in scen_rows:
    put(ws, f"A{r}", label, BOLD)
    put(ws, f"B{r}", f"='情景假设'!${ac}$35", GREEN, fmt=FM_PX)
    put(ws, f"C{r}", dcf_f, GREEN, fmt=FM_PX)
    put(ws, f"D{r}", sotp_f, GREEN, fmt=FM_PX)
    c = put(ws, f"E{r}", f"=AVERAGE({_blend_rng.format(r=r)})", BLACK, fmt=FM_PX)
    c.fill = GREENFILL if r != 24 else REDFILL
    c.font = Font(name="Arial", bold=True, size=10)
    put(ws, f"F{r}", f"=E{r}/$B$4-1", BLACK, fmt=FM_PCT)
    put(ws, f"G{r}", f"=E{r}*(1+'情景假设'!${ac}$10)", BLACK, fmt=FM_PX)
    r += 1
from openpyxl.formatting.rule import CellIsRule  # noqa: E402

ws.conditional_formatting.add(
    "F22:F24", CellIsRule(operator="lessThan", formula=["0"],
                          font=Font(name="Arial", color="CC0000", bold=True)))
ws.conditional_formatting.add(
    "F22:F24", CellIsRule(operator="greaterThan", formula=["0"],
                          font=Font(name="Arial", color="008000", bold=True)))
_note25 = "注：一年后目标价 = 综合目标价 × (1+该情景 WACC)，即公允价值按要求回报率滚动一年（未扣股息）"
if not _sotp_in:
    _note25 += f"；SOTP 不入综合（主分部利润占比 {d['meta']['seg1_share']:.0%} >= 85%，与 PE 法为同一笔盈利）"
put(ws, "A25", _note25, GREEN, border=False)
_mth = "三法" if _sotp_in else "PE/DCF 两法"
_spread = " / ".join(f"{sc} {d['scenarios'][sc].get('method_spread') or '—'}x"
                     for sc in ("bear", "base", "bull"))
put(ws, "A26", f"方法离散度（同情景{_mth} max/min）：{_spread}——离散大说明各法分歧大，"
               "综合目标价可信度降低，应分别看各法并参考反向 DCF", GREEN, border=False)

# 价值交易区间：历史已实现 NTM PE 分位 × base 前瞻 EPS——与上方三情景互为对照
# （情景=基本面情景各自的公允价，bear 是 EPS↓×PE↓ 双压；这块回答「按该票自己的
# 历史倍数分布，base 盈利下会交易在哪个区间」）。PE 分位是引擎从 facts.pe_band
# 取的常量，价格行/距现价行是活公式（改 base EPS 或现价即联动）。
# 布局契约：块从 27 行起、价格行=起始行+3。verify_report 的检查列按同一 _qs 规则
# 动态推导（不再写死 D30），但行号仍以此为契约——改起始行须两边同步。
_row = 27
_tr = d.get("trading_range")
if _tr:
    put(ws, f"A{_row}", f"价值交易区间（{_tr['window']}已实现 NTM PE 分位 × base 前瞻 EPS）",
        H2, border=False)
    _row += 1
    _qs = [q for q in ("10", "25", "50", "75", "90") if q in _tr["pe"]]
    put(ws, f"A{_row}", "历史 PE 分位", BOLD, fill=HDRFILL)
    for j, q in enumerate(_qs):
        put(ws, f"{get_column_letter(2+j)}{_row}", f"P{q}", BOLD, fill=HDRFILL)
    _row += 1
    put(ws, f"A{_row}", "PE (×)", BOLD)
    for j, q in enumerate(_qs):
        put(ws, f"{get_column_letter(2+j)}{_row}", _tr["pe"][q], BLUE, fmt=FM_X)
    _row += 1
    put(ws, f"A{_row}", "× base 前瞻 EPS", BOLD)
    for j in range(len(_qs)):
        col = get_column_letter(2 + j)
        put(ws, f"{col}{_row}", f"={col}{_row-1}*'情景假设'!$C$34", BLACK, fmt=FM_PX)
    _row += 1
    put(ws, f"A{_row}", "距现价", BOLD)
    for j in range(len(_qs)):
        col = get_column_letter(2 + j)
        put(ws, f"{col}{_row}", f"={col}{_row-1}/$B$4-1", BLACK, fmt=FM_PCT)
    _row += 1
    put(ws, f"A{_row}", f"注：分位取自 {_tr['days']} 个交易日的逐日已实现 NTM PE"
                        f"（basis={_tr['basis']}，一次性畸变窗口已双侧剔除"
                        + (f"；5年全窗中位 {_tr['full_window_p50']}x 供对照"
                           if _tr.get("full_window_p50") else "")
                        + "）。P25~P75 为核心区间——「按历史倍数该在哪交易」，"
                          "与上方情景估值（基本面公允价）分歧本身即是信号。",
        GREEN, border=False)
    _row += 2

# v2 诊断红旗区：engine diagnostics 的 red/yellow 警告逐条落进摘要——
# 警告和 headline 数字必须出现在同一张表上，不能只活在引擎 stdout 里
_flags = [(sc, lv, msg) for sc in ("bear", "base", "bull")
          for lv, msg in d["scenarios"][sc].get("warnings", [])]
if _flags:
    put(ws, f"A{_row}", "合理性红旗（engine v2 diagnostics——red 项判断层已复审一次，仍越界则代表其坚持该假设）：",
        Font(name="Arial", color="CC0000", bold=True, size=10), border=False)
    _row += 1
    for sc, lv, msg in _flags:
        put(ws, f"A{_row}", f"{'⛔' if lv == 'red' else '⚠'} [{sc}] {msg}",
            Font(name="Arial", color="CC0000" if lv == "red" else "B8860B", size=10),
            border=False)
        _row += 1

put(ws, f"A{_row}", "判断层注记（均有财报原文出处）：", H2, border=False)
for j, note in enumerate(d["notes"]):
    put(ws, f"A{_row+1+j}", note, GREEN, border=False)

# ============================================================ 出处
ws = wb.create_sheet("出处")
ws.column_dimensions["A"].width = 30
ws.column_dimensions["B"].width = 105
put(ws, "A1", "数据出处与调整口径", TITLE, border=False)
sources = [
    ("下载器 manifest", d["manifest"].strip()),
    ("XBRL companyfacts", f"https://data.sec.gov/api/xbrl/companyfacts/CIK*.json（{T}）—— 营收/营业利润/净利/EPS/CFO/Capex/现金/债务/股本"),
    ("现价", f"yfinance {T} fast_info，{d['date']}：${d['meta']['price']}，市值 ${d['meta']['mcap']:,}M"),
    ("TTM 口径", "损益类=最近四个离散季度加总（Q4=年度-前三季）；现金流类=最新年度+本财年YTD-上年同期YTD（10-Q现金流表为累计口径）"),
    ("调整后净利", d["adj_note"]),
    ("净现金", d["net_cash_note"]),
    ("估值语义版本", (f"semantics_version={d.get('semantics_version', 1)}"
                     + ("（v3, 2026-08-14：v2 之上 base 目标 PE 默认锚历史 NTM 带"
                        "近3年子窗 P50、偏离 ±15% 须给证据；"
                        f"config 声明版本 v{d.get('config_semantics_version', 1)}；"
                        "锚前(≤v2)与锚后(v3) 的倍数假设与目标价水平不可直接对比）"
                        if d.get("semantics_version", 1) >= 3 else
                        "（v2, 2026-07-22：跨情景一致性规则 / SOTP 降级 / 诊断红旗；"
                        f"config 声明版本 v{d.get('config_semantics_version', 1)}；"
                        "v1 与 v2 的倍数假设与综合口径不可直接对比）"
                        if d.get("semantics_version", 1) >= 2 else
                        "（v1：情景盈利×情景倍数，三法均值，无一致性联动）"))),
    ("免责声明", "本报告为 sec-filing-downloader 项目的数据分析输出，所有估值结果由假设驱动，不构成任何投资建议。"),
]
r = 3
for label, note in sources:
    put(ws, f"A{r}", label, BOLD)
    put(ws, f"B{r}", note, BLACK, wrap=True)
    r += 1

os.makedirs(os.path.dirname(OUT) or ".", exist_ok=True)
wb.save(OUT)
print("saved:", OUT)
